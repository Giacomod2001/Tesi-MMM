"""
Cruscotto Excel direzionale (MMM Budget Optimizer).

Workbook formattato per la lettura manageriale, con grafici nativi Excel e
formattazione condizionale (non immagini). Fogli:

  Sintesi_Canali       KPI + allocazione per CANALE + 4 grafici grandi
  Sintesi_Campagne     idem, per campagna
  Pianificazione_Mensile  piano 12 mesi (stagionalita' a colori) + grafico combo
  Metriche_Storiche    storico per campagna + grafico spesa

Entry point: build_dashboard(...).
"""
from __future__ import annotations

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import quarter as Q

# --------------------------------------------------------------- stile
NAVY, BLUE, LIGHT, BAND = "1F4E79", "2E75B6", "DDEBF7", "F2F7FC"
GREEN, GREEN_T = "C6EFCE", "006100"
RED, RED_T = "FFC7CE", "9C0006"
YELLOW, YELLOW_T = "FFEB9C", "9C6500"
GREY_S, BLUE_S, GREEN_S = "A6A6A6", "2E75B6", "4C9F70"
WHITE = "FFFFFF"
_TH = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
EUR, PCT, NUM, EUR2 = "#,##0 €", "0.0%", "#,##0", "#,##0.00 €"
MESI = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio",
        "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = Font(b=True, size=16, color=WHITE); c.fill = _fill(NAVY)
    c.alignment = CENTER; ws.row_dimensions[1].height = 30


def _band(ws, text, ncols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = Font(b=True, color=WHITE, size=11); c.fill = _fill(BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def _headers(ws, headers, row):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.font = Font(b=True, color=WHITE); c.fill = _fill(BLUE)
        c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 30


def _series_colors(chart, colors):
    for s, col in zip(chart.series, colors):
        s.graphicalProperties.solidFill = col
        s.graphicalProperties.line.solidFill = col


# --------------------------------------------------------------- tabelle dati
def _metrics(name, bud, sp, cs, ca, sat):
    return {
        name: None, "Budget Consigliato": bud, "Mix %": 0.0, "Spesa Storica": sp,
        "Var. Budget EUR": bud - sp, "Var. Budget %": (bud / sp - 1) if sp > 0 else 0.0,
        "Conv Storiche": cs, "Conv Attese": ca,
        "Var. Conv %": (ca / cs - 1) if cs > 0 else 0.0,
        "CPA Storico": sp / cs if cs > 0 else 0.0,
        "CPA Previsto": bud / ca if ca > 0 else 0.0,
        "Var. CPA %": ((bud / ca) / (sp / cs) - 1) if cs > 0 and ca > 0 and sp > 0 else 0.0,
        "Saturazione": "ALTO" if sat > 0.6 else ("MEDIO" if sat > 0.3 else "BASSO"),
    }


def _channel_table(canali, curves) -> pd.DataFrame:
    df = canali[canali["canale"] != "TOTALE"]
    rows = []
    for _, r in df.iterrows():
        ch = r["canale"]; c = curves[ch]
        bud = float(r["spesa consigliata (EUR)"])
        sat = Q._steady_response(bud / Q.WEEKS, c["lam"], c["ec"], c["slope"], c["scale"])
        d = _metrics("Canale", bud, float(r["spesa attuale (EUR)"]),
                     float(r["candidature ora"]), float(r["candidature attese"]), sat)
        d["Canale"] = ch.upper()
        rows.append(d)
    t = pd.DataFrame(rows)
    t["Mix %"] = t["Budget Consigliato"] / t["Budget Consigliato"].sum()
    return t.sort_values("Budget Consigliato", ascending=False)


def _campaign_table(canali, campagne, curves) -> pd.DataFrame:
    cand_now_ch = dict(zip(canali["canale"], canali["candidature ora"]))
    rows = []
    for _, r in campagne.iterrows():
        ch = r["canale"]; c = curves[ch]
        bud = float(r["budget consigliato (EUR)"])
        cs = float(cand_now_ch.get(ch, 0.0)) * float(r["quota attuale"])
        sat = Q._steady_response(bud / Q.WEEKS, c["lam"], c["ec"], c["slope"], c["scale"])
        d = _metrics("Campagna", bud, float(r["spesa attuale (EUR)"]),
                     cs, float(r["candidature attese"]), sat)
        d["Campagna"] = f"{ch.upper()}_{r['campagna'].upper()}"
        rows.append(d)
    t = pd.DataFrame(rows)
    t["Mix %"] = t["Budget Consigliato"] / t["Budget Consigliato"].sum()
    return t.sort_values("Budget Consigliato", ascending=False)


# --------------------------------------------------------------- foglio allocazione
def _alloc_sheet(ws, title, tab):
    ws.sheet_view.showGridLines = False
    cols = list(tab.columns); ncol = len(cols)
    ci = {h: i + 1 for i, h in enumerate(cols)}
    _title(ws, title, ncol)

    tot_bud = tab["Budget Consigliato"].sum(); tot_sp = tab["Spesa Storica"].sum()
    tot_cs = tab["Conv Storiche"].sum(); tot_ca = tab["Conv Attese"].sum()
    kpis = [("INVESTIMENTO SETT.", tot_bud / Q.WEEKS, EUR),
            ("INVESTIMENTO MENSILE", tot_bud / 3.0, EUR),
            ("CONV. ATTESE / SETT.", tot_ca / Q.WEEKS, NUM),
            ("CPA MEDIO PREVISTO", tot_bud / max(tot_ca, 1), EUR2)]
    span = max(ncol // 4, 2)
    for i, (lab, val, fmt) in enumerate(kpis):
        c0 = 1 + i * span; c1 = c0 + span - 1
        ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c1)
        lc = ws.cell(3, c0, lab); lc.fill = _fill(BLUE); lc.alignment = CENTER
        lc.font = Font(b=True, color=WHITE, size=10)
        ws.merge_cells(start_row=4, start_column=c0, end_row=5, end_column=c1)
        vc = ws.cell(4, c0, val); vc.fill = _fill(LIGHT); vc.number_format = fmt
        vc.font = Font(b=True, color=NAVY, size=18); vc.alignment = CENTER
    ws.row_dimensions[4].height = 24

    band, hdr = 7, 8
    _band(ws, "ANALISI DI ALLOCAZIONE E VARIAZIONI", ncol, band)
    _headers(ws, cols, hdr)
    fmts = {"Budget Consigliato": EUR, "Mix %": PCT, "Spesa Storica": EUR,
            "Var. Budget EUR": EUR, "Var. Budget %": PCT, "Conv Storiche": NUM,
            "Conv Attese": NUM, "Var. Conv %": PCT, "CPA Storico": EUR2,
            "CPA Previsto": EUR2, "Var. CPA %": PCT}
    sat_c = {"ALTO": (RED, RED_T), "MEDIO": (YELLOW, YELLOW_T), "BASSO": (GREEN, GREEN_T)}
    for i, (_, row) in enumerate(tab.iterrows()):
        rr = hdr + 1 + i
        for j, col in enumerate(cols, 1):
            c = ws.cell(rr, j, row[col]); c.border = BORDER
            if i % 2:
                c.fill = _fill(BAND)
            if col in fmts:
                c.number_format = fmts[col]
            if j == 1:
                c.font = Font(b=True)
            if col == "Saturazione":
                fg, tx = sat_c[row[col]]; c.fill = _fill(fg)
                c.font = Font(b=True, color=tx); c.alignment = Alignment(horizontal="center")
            elif col in ("Var. Conv %", "Var. Budget %"):
                c.font = Font(color=GREEN_T if row[col] >= 0 else RED_T)
            elif col == "Var. CPA %":
                c.font = Font(color=GREEN_T if row[col] <= 0 else RED_T)
    n = len(tab); last = hdr + n; tr = last + 1
    for j in range(1, ncol + 1):
        cc = ws.cell(tr, j); cc.fill = _fill(NAVY); cc.font = Font(b=True, color=WHITE)
    ws.cell(tr, 1, "TOTALE")
    for col, val, fmt in (("Budget Consigliato", tot_bud, EUR), ("Spesa Storica", tot_sp, EUR),
                          ("Var. Budget EUR", tot_bud - tot_sp, EUR),
                          ("Conv Storiche", tot_cs, NUM), ("Conv Attese", tot_ca, NUM)):
        c = ws.cell(tr, ci[col], val); c.number_format = fmt

    widths = [26, 15, 8, 13, 14, 11, 12, 12, 11, 12, 12, 11, 12]
    for j, w in enumerate(widths[:ncol], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "B9"
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(ncol)}{last}"

    # ---- grafici grandi, sotto la tabella (orizzontali, etichette, colori)
    cats = Reference(ws, min_col=1, min_row=hdr + 1, max_row=last)
    cr = tr + 2

    do = DoughnutChart(); do.title = "MIX BUDGET (% del totale)"
    do.height = 9.5; do.width = 17; do.style = 26
    do.add_data(Reference(ws, min_col=ci["Budget Consigliato"], min_row=hdr, max_row=last),
                titles_from_data=True)
    do.set_categories(cats)
    do.dataLabels = DataLabelList(); do.dataLabels.showPercent = True
    do.dataLabels.showCatName = True; do.dataLabels.numFmt = "0%"
    ws.add_chart(do, f"A{cr}")

    def hbar(title, c1, c2, colors, fmt):
        ch = BarChart(); ch.type = "bar"; ch.title = title
        ch.height = 9.5; ch.width = 17; ch.gapWidth = 60; ch.style = 10
        ch.add_data(Reference(ws, min_col=c1, min_row=hdr, max_row=last), titles_from_data=True)
        ch.add_data(Reference(ws, min_col=c2, min_row=hdr, max_row=last), titles_from_data=True)
        ch.set_categories(cats); _series_colors(ch, colors)
        ch.dataLabels = DataLabelList(); ch.dataLabels.showVal = True
        ch.dataLabels.numFmt = fmt
        ch.legend.position = "b"
        return ch

    ws.add_chart(hbar("CONFRONTO SPESA: storica vs consigliata",
                      ci["Spesa Storica"], ci["Budget Consigliato"],
                      [GREY_S, BLUE_S], "#,##0"), f"J{cr}")
    ws.add_chart(hbar("CPA: storico vs previsto",
                      ci["CPA Storico"], ci["CPA Previsto"],
                      [GREY_S, BLUE_S], "#,##0"), f"A{cr + 20}")
    ws.add_chart(hbar("CONVERSIONI: storiche vs attese",
                      ci["Conv Storiche"], ci["Conv Attese"],
                      [GREY_S, GREEN_S], "#,##0"), f"J{cr + 20}")


# --------------------------------------------------------------- mensile / storico
def _sheet_mensile(wb, tot_bud, tot_conv_att, seas):
    ws = wb.create_sheet("Pianificazione_Mensile")
    ws.sheet_view.showGridLines = False
    _title(ws, "Piano Operativo Mensile", 5)
    _headers(ws, ["Mese", "Indice Stagionalita", "Azione Strategica",
                  "Budget Suggerito", "Conv. Stimate"], 3)
    s = seas.copy()
    if "region" in s.columns:
        s = s[s["region"].astype(str) == "*"]
    s["m"] = pd.to_datetime(s["week"]).dt.month
    idx = s.groupby("m")["seasonal_index"].mean()
    idx = (idx / idx.mean()).reindex(range(1, 13)).fillna(1.0)
    avg = tot_bud / 3.0; cpe = tot_conv_att / max(tot_bud, 1)
    for i, m in enumerate(range(1, 13)):
        rr = 4 + i; v = float(idx[m]); budget = avg * v
        az = "PUSH" if v > 1.03 else ("RIDUCI" if v < 0.97 else "MANTIENI")
        for j, x in enumerate([MESI[i], round(v, 2), az, budget, budget * cpe], 1):
            c = ws.cell(rr, j, x); c.border = BORDER
            if i % 2:
                c.fill = _fill(BAND)
        ws.cell(rr, 1).font = Font(b=True)
        ws.cell(rr, 3).font = Font(b=True); ws.cell(rr, 3).alignment = Alignment(horizontal="center")
        ws.cell(rr, 4).number_format = EUR; ws.cell(rr, 5).number_format = NUM
    ws.conditional_formatting.add("B4:B15", ColorScaleRule(
        start_type="min", start_color="F8696B", mid_type="percentile",
        mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    for j, w in enumerate([14, 18, 18, 18, 16], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    bar = BarChart(); bar.type = "col"; bar.title = "BUDGET E STAGIONALITA' PER MESE"
    bar.height = 10; bar.width = 24; bar.style = 10
    bar.add_data(Reference(ws, min_col=4, min_row=3, max_row=15), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=4, max_row=15))
    _series_colors(bar, [BLUE_S])
    line = LineChart()
    line.add_data(Reference(ws, min_col=2, min_row=3, max_row=15), titles_from_data=True)
    line.y_axis.axId = 200; line.y_axis.title = "indice"
    _series_colors(line, [RED_T])
    bar.y_axis.title = "budget (EUR)"; bar.legend.position = "b"; bar += line
    ws.add_chart(bar, "A18")


def _sheet_storico(wb, media):
    ws = wb.create_sheet("Metriche_Storiche")
    ws.sheet_view.showGridLines = False
    _title(ws, "Dettaglio Storico Canali", 5)
    _headers(ws, ["Campagna", "Spend Tot EUR", "Spend Medio Sett",
                  "Click Medio Sett", "CPA Medio Storico"], 3)
    nw = media["week"].nunique()
    g = media.groupby(["channel", "campaign"]).agg(
        spend=("spend", "sum"), clicks=("clicks", "sum"),
        conv=("platform_conversions", "sum")).reset_index()
    g["name"] = g["channel"].str.upper() + "_" + g["campaign"].str.upper()
    g = g.sort_values("spend", ascending=False)
    for i, (_, r) in enumerate(g.iterrows()):
        rr = 4 + i
        cpa = r["spend"] / r["conv"] if r["conv"] > 0 else 0.0
        for j, x in enumerate([r["name"], r["spend"], r["spend"] / nw,
                               r["clicks"] / nw, cpa], 1):
            c = ws.cell(rr, j, x); c.border = BORDER
            if i % 2:
                c.fill = _fill(BAND)
        ws.cell(rr, 2).number_format = EUR; ws.cell(rr, 3).number_format = EUR
        ws.cell(rr, 4).number_format = NUM; ws.cell(rr, 5).number_format = EUR2
    last = 3 + len(g)
    for j, w in enumerate([28, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    bar = BarChart(); bar.type = "bar"; bar.title = "SPESA TOTALE PER CAMPAGNA"
    bar.height = 11; bar.width = 18; bar.style = 10
    bar.add_data(Reference(ws, min_col=2, min_row=3, max_row=last), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=4, max_row=last))
    _series_colors(bar, [BLUE_S]); bar.legend = None
    ws.add_chart(bar, "G3")


# --------------------------------------------------------------- entry point
def build_dashboard(canali, campagne, summary, media, seas, rev_per_conv, path) -> str:
    nw = media["week"].nunique()
    hist = (media.groupby("channel")["spend"].sum() / nw).to_dict()
    curves = Q.build_curves(summary, hist)
    ch_tab = _channel_table(canali, curves)
    cp_tab = _campaign_table(canali, campagne, curves)

    wb = Workbook()
    _alloc_sheet(wb.active, "MMM BUDGET OPTIMIZER  -  Cruscotto per CANALE", ch_tab)
    wb.active.title = "Sintesi_Canali"
    _alloc_sheet(wb.create_sheet("Sintesi_Campagne"),
                 "MMM BUDGET OPTIMIZER  -  Cruscotto per CAMPAGNA", cp_tab)
    _sheet_mensile(wb, float(ch_tab["Budget Consigliato"].sum()),
                   float(ch_tab["Conv Attese"].sum()), seas)
    _sheet_storico(wb, media)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return path
